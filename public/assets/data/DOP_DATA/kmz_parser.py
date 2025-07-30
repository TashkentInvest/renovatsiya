#!/usr/bin/env python3
"""
KMZ to XLSX Data Extractor - Fixed Structure
Extracts polygon coordinates and descriptions from KMZ files and exports to XLSX
"""

import zipfile
import xml.etree.ElementTree as ET
import pandas as pd
import re
from pathlib import Path
import argparse
import sys
from typing import List, Dict, Tuple, Any
import traceback

def extract_kml_from_kmz(kmz_path: str) -> str:
    """Extract KML content from KMZ file"""
    try:
        with zipfile.ZipFile(kmz_path, 'r') as kmz:
            # Look for KML files
            kml_files = [f for f in kmz.namelist() if f.endswith('.kml')]

            if not kml_files:
                raise ValueError("No KML file found in KMZ archive")

            # Use the first KML file (usually doc.kml)
            kml_file = kml_files[0]
            print(f"📄 Reading KML file: {kml_file}")

            with kmz.open(kml_file) as kml:
                content = kml.read().decode('utf-8')
                print(f"📏 KML content size: {len(content)} characters")
                return content

    except Exception as e:
        print(f"❌ Error extracting KML from KMZ: {e}")
        raise

def clean_text(text: str) -> str:
    """Clean and normalize text content"""
    if not text:
        return ""

    # Remove HTML tags
    text = re.sub(r'<[^>]+>', '', text)
    # Normalize whitespace
    text = re.sub(r'\s+', ' ', text)
    # Remove special characters that might cause issues
    text = text.strip()

    return text

def parse_coordinates(coord_string: str) -> List[Tuple[float, float]]:
    """Parse coordinate string from KML and return list of (lon, lat) tuples"""
    if not coord_string:
        return []

    coordinates = []
    coord_string = coord_string.strip()

    try:
        # Remove any extra whitespace and normalize
        coord_string = re.sub(r'\s+', ' ', coord_string)

        # Try different splitting methods
        points = []

        # Method 1: Split by whitespace
        if ' ' in coord_string:
            points = coord_string.split()
        # Method 2: Split by newlines
        elif '\n' in coord_string:
            points = coord_string.split('\n')
        # Method 3: Single coordinate string
        else:
            points = [coord_string]

        for point in points:
            point = point.strip()
            if not point:
                continue

            # Each point should be "longitude,latitude" or "longitude,latitude,altitude"
            coords = point.split(',')
            if len(coords) >= 2:
                try:
                    lon = float(coords[0].strip())
                    lat = float(coords[1].strip())

                    # Basic validation for reasonable coordinates
                    if -180 <= lon <= 180 and -90 <= lat <= 90:
                        coordinates.append((lon, lat))
                except (ValueError, IndexError):
                    continue

    except Exception as e:
        print(f"⚠️  Warning: Error parsing coordinates: {e}")
        print(f"   Coordinate string preview: {coord_string[:100]}...")

    return coordinates

def parse_description(description: str) -> Dict[str, str]:
    """Parse description field and extract structured data"""
    if not description:
        return {}

    data = {}
    description = clean_text(description)

    # Split by lines and parse key-value pairs
    lines = description.split('\n')

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Try different separators
        separators = [' - ', ':', '-', '=', '–', '—']

        for sep in separators:
            if sep in line:
                parts = line.split(sep, 1)
                if len(parts) == 2:
                    key = parts[0].strip()
                    value = parts[1].strip()

                    if key and value:
                        # Clean up quotes and extra spaces
                        value = value.replace('""', '"').strip()
                        data[key] = value
                        break

    return data

def find_all_elements_by_tag(root, tag_name: str) -> List:
    """Recursively find all elements with a specific tag name (ignoring namespace)"""
    elements = []

    def recursive_search(element):
        # Check if current element matches (ignore namespace)
        if element.tag.endswith(tag_name) or element.tag == tag_name:
            elements.append(element)

        # Search children
        for child in element:
            recursive_search(child)

    recursive_search(root)
    return elements

def extract_polygons_from_kml(kml_content: str) -> List[Dict[str, Any]]:
    """Extract all polygon data from KML content"""
    try:
        # Clean content and handle BOM
        if kml_content.startswith('\ufeff'):
            kml_content = kml_content[1:]

        # Parse XML
        root = ET.fromstring(kml_content)
        print(f"📊 Root element: {root.tag}")

        polygons = []

        # Find all Placemark elements
        placemarks = find_all_elements_by_tag(root, 'Placemark')
        print(f"🔍 Found {len(placemarks)} placemark elements")

        for i, placemark in enumerate(placemarks):
            try:
                polygon_data = {'id': i + 1}

                # Extract name
                name_elements = find_all_elements_by_tag(placemark, 'name')
                if name_elements and name_elements[0].text:
                    polygon_data['name'] = clean_text(name_elements[0].text)
                else:
                    polygon_data['name'] = f'Polygon_{i+1}'

                # Extract description
                desc_elements = find_all_elements_by_tag(placemark, 'description')
                description = ''
                if desc_elements and desc_elements[0].text:
                    description = desc_elements[0].text

                polygon_data['raw_description'] = description
                polygon_data['parsed_data'] = parse_description(description)

                # Find coordinates from any geometry type
                coordinates = []
                coord_elements = find_all_elements_by_tag(placemark, 'coordinates')

                for coord_elem in coord_elements:
                    if coord_elem.text:
                        coords = parse_coordinates(coord_elem.text)
                        coordinates.extend(coords)

                polygon_data['coordinates'] = coordinates
                polygon_data['coordinate_count'] = len(coordinates)

                # Calculate center point
                if coordinates:
                    center_lon = sum(coord[0] for coord in coordinates) / len(coordinates)
                    center_lat = sum(coord[1] for coord in coordinates) / len(coordinates)
                    polygon_data['center_longitude'] = round(center_lon, 8)
                    polygon_data['center_latitude'] = round(center_lat, 8)
                else:
                    polygon_data['center_longitude'] = None
                    polygon_data['center_latitude'] = None

                polygons.append(polygon_data)

                # Progress indicator
                if (i + 1) % 50 == 0:
                    print(f"   Processed {i + 1} placemarks...")

            except Exception as e:
                print(f"⚠️  Warning: Error processing placemark {i+1}: {e}")
                continue

        print(f"✅ Successfully processed {len(polygons)} polygons")
        return polygons

    except ET.ParseError as e:
        print(f"❌ XML Parse Error: {e}")
        raise
    except Exception as e:
        print(f"❌ Error in extract_polygons_from_kml: {e}")
        traceback.print_exc()
        raise

def create_xlsx_output(polygons: List[Dict[str, Any]], output_path: str):
    """Create beautiful XLSX file with extracted data"""

    print(f"📝 Creating beautiful XLSX output with {len(polygons)} polygons...")

    # Prepare main data with Uzbek Cyrillic columns
    rows = []

    for polygon in polygons:
        # Get first coordinate point (one corner of polygon)
        coordinates = polygon.get('coordinates', [])

        # Round coordinates to 2 decimal places for cleaner display
        first_coord = coordinates[0] if coordinates else (None, None)

        # Create Google Maps link if coordinates exist
        google_maps_link = ""
        lat_display = ""
        lon_display = ""

        if first_coord[0] is not None and first_coord[1] is not None:
            lat_display = round(first_coord[1], 2)  # Round to 2 decimal places
            lon_display = round(first_coord[0], 2)  # Round to 2 decimal places
            google_maps_link = f"https://www.google.com/maps?q={first_coord[1]},{first_coord[0]}"

        # Base row structure with Uzbek Cyrillic columns
        row = {
            'ИД': polygon.get('id', ''),
            'Объект номи': polygon.get('name', ''),
            'Туман': '',
            'МФЙ': '',
            'Майдони (га)': '',
            'ККУ сони': '',
            'Қурилган йили': '',
            'Квартира сони': '',
            'Қурилиш ости майдони': '',
            'Бино майдони': '',
            'Нотурар жой майдони': '',
            'МКД сони': '',
            'Умумий майдон': '',
            'Турар жой майдони': '',
            'Нотурар майдони': '',
            'Турар жойи тури': '',
            'Нотурар жойи тури': '',
            'ИЖС': '',
            'ККУ (4 қаватли)': '',
            'Координата (кенглик)': lat_display,
            'Координата (узунлик)': lon_display,
            'Google Maps ҳавола': google_maps_link,
            'Жами координаталар сони': polygon.get('coordinate_count', 0),
            'Таърифи': polygon.get('raw_description', '')
        }

        # Map parsed description data to appropriate columns
        parsed_data = polygon.get('parsed_data', {})

        # Field mappings - mapping description keys to Excel columns
        field_mappings = {
            'Туман': ['туман'],
            'МФЙ': ['мфй'],
            'Майдони (га)': ['майдони'],
            'ККУ сони': ['кку сони'],
            'Қурилган йили': ['қурилган йили', 'курилган йили'],
            'Квартира сони': ['квартира сони'],
            'Қурилиш ости майдони': ['қурилиш ости', 'курилиш ости'],
            'Бино майдони': ['бино майдони'],
            'Турар жой майдони': ['турар майдони'],
            'Нотурар майдони': ['нотурар майдони'],
            'Нотурар жой майдони': ['нотурар'],
        }

        # Direct mapping from parsed data
        for parsed_key, parsed_value in parsed_data.items():
            parsed_key_clean = parsed_key.strip().lower()

            # Direct matches
            if parsed_key_clean == 'туман':
                row['Туман'] = parsed_value
            elif parsed_key_clean == 'мфй':
                row['МФЙ'] = parsed_value
            elif parsed_key_clean == 'майдони':
                # Extract just the number and unit
                if 'га' in parsed_value:
                    row['Майдони (га)'] = parsed_value.replace('га', '').strip()
                else:
                    row['Майдони (га)'] = parsed_value
            elif parsed_key_clean == 'кку сони':
                row['ККУ сони'] = parsed_value
            elif 'курилган йили' in parsed_key_clean or 'қурилган йили' in parsed_key_clean:
                row['Қурилган йили'] = parsed_value
            elif parsed_key_clean == 'квартира сони':
                row['Квартира сони'] = parsed_value
            elif 'курилиш ости' in parsed_key_clean or 'қурилиш ости' in parsed_key_clean:
                row['Қурилиш ости майдони'] = parsed_value
            elif parsed_key_clean == 'бино майдони':
                row['Бино майдони'] = parsed_value
            elif 'турар майдони' in parsed_key_clean:
                row['Турар жой майдони'] = parsed_value
            elif 'нотурар майдони' in parsed_key_clean:
                row['Нотурар майдони'] = parsed_value
            elif parsed_key_clean == 'нотурар':
                row['Нотурар жой майдони'] = parsed_value

        rows.append(row)

    # Create beautiful Excel file
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.utils import get_column_letter

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Полигонлар маълумотлари"

        # Create DataFrame
        df = pd.DataFrame(rows)

        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Style definitions
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3685', end_color='1E3685', fill_type='solid')
        data_font = Font(name='Arial', size=10)

        # Border styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Header styling
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Data styling
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

                # Special formatting for Google Maps links
                col_letter = get_column_letter(cell.column)
                if col_letter == 'V':  # Google Maps column
                    if cell.value and cell.value.startswith('https://'):
                        cell.font = Font(name='Arial', size=10, color='0000FF', underline='single')

        # Auto-adjust column widths
        column_widths = [
            ('A', 8),   # ИД
            ('B', 25),  # Объект номи
            ('C', 15),  # Туман
            ('D', 20),  # МФЙ
            ('E', 12),  # Майдони
            ('F', 10),  # ККУ сони
            ('G', 12),  # Қурилган йили
            ('H', 12),  # Квартира сони
            ('I', 15),  # Қурилиш ости
            ('J', 12),  # Бино майдони
            ('K', 15),  # Нотурар жой
            ('L', 10),  # МКД сони
            ('M', 12),  # Умумий майдон
            ('N', 15),  # Турар жой
            ('O', 15),  # Нотурар майдони
            ('P', 15),  # Турар жойи тури
            ('Q', 15),  # Нотурар жойи тури
            ('R', 8),   # ИЖС
            ('S', 12),  # ККУ 4 қаватли
            ('T', 15),  # Кенглик
            ('U', 15),  # Узунлик
            ('V', 40),  # Google Maps
            ('W', 10),  # Координаталар сони
            ('X', 30)   # Таърифи
        ]

        for col_letter, width in column_widths:
            if col_letter <= get_column_letter(ws.max_column):
                ws.column_dimensions[col_letter].width = width

        # Set row height for header
        ws.row_dimensions[1].height = 30

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add auto filter with proper range
        if ws.max_row > 1 and ws.max_column > 0:
            max_col_letter = get_column_letter(ws.max_column)
            filter_range = f"A1:{max_col_letter}{ws.max_row}"
            ws.auto_filter.ref = filter_range

        # Save workbook
        wb.save(output_path)

        print(f"   ✅ Beautiful XLSX created with {len(df)} rows")
        print(f"   🎨 Applied professional styling and formatting")
        print(f"   🔗 Added {len([r for r in rows if r['Google Maps ҳавола']])} Google Maps links")

    except ImportError:
        # Fallback to basic pandas Excel writer if openpyxl advanced features not available
        print("   ⚠️  Advanced styling not available, creating basic Excel file...")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df = pd.DataFrame(rows)
            df.to_excel(writer, sheet_name='Полигонлар маълумотлари', index=False)
        print(f"   ✅ Basic XLSX created with {len(rows)} rows")

    except Exception as e:
        print(f"❌ Error creating XLSX file: {e}")
        raise

def main():
    parser = argparse.ArgumentParser(description='Extract polygon data from KMZ file to XLSX')
    parser.add_argument('kmz_file', nargs='?', help='Path to the KMZ file')
    parser.add_argument('-o', '--output', help='Output XLSX file path',
                       default='extracted_polygons.xlsx')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose output')

    args = parser.parse_args()

    # If no file specified, use default
    if not args.kmz_file:
        default_kmz = "ALL_RENOVATION_AREA_368_303_230525.kmz"
        if Path(default_kmz).exists():
            args.kmz_file = default_kmz
            print(f"🔍 Using default KMZ file: {default_kmz}")
        else:
            print("❌ No KMZ file specified and default file not found.")
            print(f"Usage: python {sys.argv[0]} <kmz_file> [-o output.xlsx] [-v]")
            print(f"Example: python {sys.argv[0]} {default_kmz}")
            sys.exit(1)

    kmz_path = Path(args.kmz_file)
    output_path = Path(args.output)

    if not kmz_path.exists():
        print(f"❌ Error: KMZ file not found: {kmz_path}")
        sys.exit(1)

    try:
        print(f"🚀 Starting KMZ extraction process...")
        print(f"📁 Input file: {kmz_path.absolute()}")
        print(f"📁 Output file: {output_path.absolute()}")
        print()

        # Extract KML content
        kml_content = extract_kml_from_kmz(str(kmz_path))

        if args.verbose:
            print(f"📏 KML content length: {len(kml_content)} characters")
            print(f"📄 KML preview (first 300 chars):")
            print(kml_content[:300])
            print("...\n")

        # Parse polygons
        polygons = extract_polygons_from_kml(kml_content)

        if not polygons:
            print("⚠️  Warning: No polygons found in the KMZ file")
            print("This might indicate:")
            print("  • Different KML structure than expected")
            print("  • File contains non-polygon geometries")
            print("  • Parsing errors occurred")

            # Save raw KML for debugging
            debug_path = kmz_path.with_suffix('.debug.kml')
            with open(debug_path, 'w', encoding='utf-8') as f:
                f.write(kml_content)
            print(f"📄 Raw KML saved for debugging: {debug_path}")
            return

        # Show sample data
        print(f"\n📋 Sample data from first 3 polygons:")
        for i, polygon in enumerate(polygons[:3], 1):
            print(f"  {i}. Name: '{polygon.get('name', 'N/A')}'")
            print(f"     Coordinates: {len(polygon.get('coordinates', []))} points")

            if polygon.get('center_longitude') and polygon.get('center_latitude'):
                print(f"     Center: {polygon['center_longitude']:.6f}, {polygon['center_latitude']:.6f}")

            parsed_data = polygon.get('parsed_data', {})
            if parsed_data:
                print(f"     Parsed fields: {len(parsed_data)} items")
                for key, value in list(parsed_data.items())[:3]:
                    print(f"       • {key}: {value}")

            if polygon.get('raw_description'):
                desc_preview = polygon['raw_description'][:100].replace('\n', ' ')
                if len(polygon['raw_description']) > 100:
                    desc_preview += "..."
                print(f"     Description: {desc_preview}")
            print()

        # Create XLSX output
        create_xlsx_output(polygons, str(output_path))

        # Summary statistics
        total_coords = sum(len(p.get('coordinates', [])) for p in polygons)
        total_parsed_fields = sum(len(p.get('parsed_data', {})) for p in polygons)

        print(f"\n🎉 Extraction completed successfully!")
        print(f"📊 Summary Statistics:")
        print(f"   • Total polygons: {len(polygons)}")
        print(f"   • Total coordinate points: {total_coords:,}")
        print(f"   • Average points per polygon: {total_coords/len(polygons):.1f}")
        print(f"   • Total parsed fields: {total_parsed_fields}")
        print(f"   • Average fields per polygon: {total_parsed_fields/len(polygons):.1f}")
        print(f"📁 Output file: {output_path.absolute()}")

    except Exception as e:
        print(f"\n❌ Fatal error: {e}")
        if args.verbose:
            print("\n🔍 Full error traceback:")
            traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
